# -*- coding: utf-8 -*-
"""56 sorunun tam listesi + citation toplamları + top kaynaklar -> questions.json"""
import csv, re, json
from collections import defaultdict

BASE = "/Users/Erdo/Desktop/Claude Projects/Özdilek/ai özdilek/"
DATA = BASE + "ozdilek-ai-visibility-report/data/"

def prompt_map(fn):
    with open(DATA + fn, encoding="utf-8-sig") as f:
        lines = [l for l in f if not l.startswith("#") or l.startswith("#,")]
    pm = defaultdict(list)
    for r in csv.DictReader(lines):
        for part in (r["Top Prompts"] or "").split(" | "):
            m = re.match(r"^(.*?)\s*\((\d+)\)\s*$", part.strip())
            if m:
                pm[m.group(1).strip()].append((r["Domain"], int(m.group(2))))
    return pm

mk = prompt_map("marka-gorunurluk-analizi.csv")
pz = prompt_map("pazaryeri-gorunurluk-analizi.csv")

# 56 soru master listesi (Excel'den)
from openpyxl import load_workbook
wb = load_workbook(BASE + "Özdilekteyim - AI Görünürlük Prompt Seti (2 Kademe).xlsx")
ws = wb["Prompt Seti (2 Kademe)"]
def comp_in(c): return any(b in c for b in ["trendyol","hepsiburada","amazon","aliexpress","temu","shein"])
qs = []
for row in ws.iter_rows(min_row=2, values_only=True):
    _, kademe, kategori, alt, prompt, *_ = row[:6]
    lens = "z" if str(kademe).startswith("Kademe 1") else "m"
    c = prompt.lower()
    if ("dışında" in c or "yerine" in c) and comp_in(c): cat = "alternatives"
    elif ("mı yoksa" in c) or ("sitesinden mi" in c and "pazaryerinden mi" in c): cat = "vs"
    elif "nereden" in c: cat = "custom"
    else: cat = "best"
    qs.append({"q": prompt.strip(), "p": lens, "c": cat, "tags": kategori})

out = []
for item in qs:
    pm = pz if item["p"] == "z" else mk
    # eşleştir (tam metin)
    entry = pm.get(item["q"])
    if entry is None:
        # yakın eşleşme dene
        for k in pm:
            if k[:40] == item["q"][:40]:
                entry = pm[k]; break
    total = sum(c for _, c in entry) if entry else 0
    ozd = sum(c for d, c in entry if "ozdilek" in d) if entry else 0
    top = sorted(entry, key=lambda x: -x[1])[:6] if entry else []
    out.append({
        "q": item["q"], "p": item["p"], "c": item["c"],
        "n": total, "o": ozd,
        "s": [[d, c] for d, c in top],
    })

out.sort(key=lambda x: (-x["n"]))
with open(DATA + "questions.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

# favicon ihtiyacı: tüm top-6 domainleri
doms = sorted({d for o in out for d, _ in o["s"]})
print("soru:", len(out), "| citationlı:", sum(1 for o in out if o["n"] > 0))
print("top-kaynak domain sayısı:", len(doms))
import os
have = {f.replace(".png", "").replace("-", ".") for f in os.listdir(DATA + "../assets/logos") if f.endswith(".png")}
def slug2dom(s): return s
missing = [d for d in doms if d.replace(".", "-") + ".png" not in os.listdir(DATA + "../assets/logos")]
print("eksik favicon:", len(missing))
print(" ".join(missing))
